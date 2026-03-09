import random
import openpyxl
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from django.db import transaction
from django.utils import timezone
from django.db import models
from .models import Exam, Question, Option, Attempt, AttemptQuestion, AttemptAnswer
from .serializers import ExamSerializer, QuestionSerializer, AttemptSerializer, AttemptAnswerSerializer


class IsStaff(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and request.user.is_staff


class ImportExamView(APIView):
    """
    Importa un examen completo desde un archivo Excel (.xlsx).

    Formato esperado (hoja única):
    - Fila 1: encabezados (fijos)
    - Filas siguientes: una fila por opción de respuesta

    Columnas requeridas:
        exam_name           | Nombre del examen (mismo valor para todo el examen)
        exam_description    | Descripción del examen (opcional, solo se lee la 1ª fila)
        questions_per_attempt | Preguntas por intento (solo fila 1, default 10)
        max_scored_attempts | Intentos puntuables (solo fila 1, default 3)
        max_points          | Puntos máximos (solo fila 1, default 100)
        question_text       | Texto de la pregunta
        question_type       | single_choice / multiple_choice / open_ended
        question_points     | Puntos de la pregunta (default 10)
        time_limit_seconds  | Tiempo límite en segundos (default 60)
        option_text         | Texto de la opción (vacío si es open_ended)
        is_correct          | TRUE / FALSE (si esta opción es correcta)

    Lógica de importación:
    - Si el examen (por nombre exacto) ya existe → se REEMPLAZAN todas sus preguntas
      (los intentos y resultados anteriores se conservan pero quedarán huérfanos).
    - Si no existe → se crea un examen nuevo.
    - El admin recibe un resumen: preguntas creadas, opciones creadas, modo (nuevo/reemplazado).
    """
    permission_classes = [IsStaff]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No se proporcionó ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj.name.endswith('.xlsx'):
            return Response({'error': 'El archivo debe ser formato .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception:
            return Response({'error': 'No se pudo leer el archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)

        # Leer encabezados
        header_row = [str(cell.value).strip().lower().replace(' ', '_') if cell.value else '' for cell in ws[1]]
        required_cols = {'exam_name', 'question_text', 'question_type', 'option_text', 'is_correct'}

        if not required_cols.issubset(set(header_row)):
            missing = required_cols - set(header_row)
            return Response(
                {'error': f'Faltan columnas requeridas: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        col = {name: header_row.index(name) for name in header_row if name}

        def get(row, name, default=''):
            if name not in col:
                return default
            val = row[col[name]]
            return val if val is not None else default

        def to_bool(val):
            return str(val).strip().upper() in ('TRUE', '1', 'SI', 'SÍ', 'YES', 'VERDADERO')

        # Leer todas las filas (saltando vacías)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            rows.append(row)

        if not rows:
            return Response({'error': 'El archivo no contiene datos.'}, status=status.HTTP_400_BAD_REQUEST)

        # Metadatos del examen (de la primera fila)
        first = rows[0]
        exam_name = str(get(first, 'exam_name')).strip()
        if not exam_name:
            return Response({'error': 'El campo exam_name no puede estar vacío.'}, status=status.HTTP_400_BAD_REQUEST)

        exam_description = str(get(first, 'exam_description', '')).strip()
        try:
            questions_per_attempt = int(get(first, 'questions_per_attempt', 10))
        except (ValueError, TypeError):
            questions_per_attempt = 10
        try:
            max_scored_attempts = int(get(first, 'max_scored_attempts', 3))
        except (ValueError, TypeError):
            max_scored_attempts = 3
        try:
            max_points = int(get(first, 'max_points', 100))
        except (ValueError, TypeError):
            max_points = 100

        # Agrupar filas por pregunta
        questions_data = []  # [{text, type, points, time, options:[{text, is_correct}]}]
        current_q = None

        for row in rows:
            q_text = str(get(row, 'question_text', '')).strip()
            q_type = str(get(row, 'question_type', 'single_choice')).strip().lower()
            opt_text = str(get(row, 'option_text', '')).strip()
            is_correct = to_bool(get(row, 'is_correct', False))

            try:
                q_points = int(get(row, 'question_points', 10))
            except (ValueError, TypeError):
                q_points = 10
            try:
                time_limit = int(get(row, 'time_limit_seconds', 60))
            except (ValueError, TypeError):
                time_limit = 60

            # Si hay un texto de pregunta y es distinto al actual, creamos una nueva pregunta
            # Si el texto de pregunta está vacío, asumimos que es una opción de la pregunta actual
            if q_text and (not current_q or q_text != current_q['text']):
                current_q = {
                    'text': q_text,
                    'type': q_type if q_type in ('single_choice', 'multiple_choice', 'open_ended') else 'single_choice',
                    'points': q_points,
                    'time_limit': time_limit,
                    'options': [],
                }
                questions_data.append(current_q)

            # Agregar opción a la pregunta actual (si existe)
            if current_q and opt_text:
                current_q['options'].append({'text': opt_text, 'is_correct': is_correct})

        if not questions_data:
            return Response({'error': 'No se encontraron preguntas en el archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        # Crear/actualizar en BD
        errores = []
        with transaction.atomic():
            exam, created = Exam.objects.get_or_create(
                name=exam_name,
                defaults={
                    'description': exam_description,
                    'questions_per_attempt': questions_per_attempt,
                    'max_scored_attempts': max_scored_attempts,
                    'max_points': max_points,
                }
            )

            if not created:
                # Actualizar metadatos (no borrar intentos, sí borrar preguntas)
                exam.description = exam_description or exam.description
                exam.questions_per_attempt = questions_per_attempt
                exam.max_scored_attempts = max_scored_attempts
                exam.max_points = max_points
                exam.save()
                # Borrar preguntas anteriores (opciones se borran en cascada)
                exam.questions.all().delete()

            total_questions = 0
            total_options = 0

            for q_data in questions_data:
                try:
                    question = Question.objects.create(
                        exam=exam,
                        text=q_data['text'],
                        question_type=q_data['type'],
                        points=q_data['points'],
                        time_limit_seconds=q_data['time_limit'],
                    )
                    total_questions += 1

                    for opt in q_data['options']:
                        Option.objects.create(
                            question=question,
                            text=opt['text'],
                            is_correct=opt['is_correct'],
                        )
                        total_options += 1

                except Exception as e:
                    errores.append(f"Pregunta '{q_data['text'][:40]}': {str(e)}")

            # Actualiza el banco de preguntas
            exam.bank_total_questions = total_questions
            exam.save()

        return Response({
            'modo': 'creado' if created else 'reemplazado',
            'exam_id': exam.id,
            'exam_name': exam.name,
            'preguntas_creadas': total_questions,
            'opciones_creadas': total_options,
            'errores': errores,
        }, status=status.HTTP_200_OK)

class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.is_staff:
            return Exam.objects.all()
        return Exam.objects.filter(is_active=True, is_enabled=True)
    
    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def toggle_enabled(self, request, pk=None):
        if not request.user.is_staff:
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        exam = self.get_object()
        exam.is_enabled = not exam.is_enabled
        exam.save()
        return Response({'is_enabled': exam.is_enabled})

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def export_csv(self, request, pk=None):
        if not request.user.is_staff:
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
            
        import csv
        from django.http import HttpResponse
        
        exam = self.get_object()
        # Get all questions for this exam to fix columns
        questions = list(Question.objects.filter(exam=exam).order_by('id'))
        attempts = Attempt.objects.filter(exam=exam, status='completed').order_by('started_at')
        
        response = HttpResponse(content_type='text/csv')
        filename = f'resultados_{exam.name.replace(" ", "_").lower()}.csv'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        # Add UTF-8 BOM for Excel compatibility with accents
        response.write('\ufeff'.encode('utf8'))
        
        writer = csv.writer(response)
        
        # Pre-fetch question options for letter mapping
        options_map = {}
        for q in questions:
            options_map[q.id] = list(Option.objects.filter(question=q).order_by('id'))
        
        # Header construction
        header = ['ID', 'Start time', 'Completion time', 'Email', 'Name', 'DNI', 'Total points', 'Nº Intento']
        for q in questions:
            header.append(q.text)
            header.append(f'Points - {q.text}')
        
        writer.writerow(header)
        
        for row_idx, obj in enumerate(attempts, start=1):
            # Basic info
            row = [
                row_idx,
                obj.started_at.strftime("%Y-%m-%d %H:%M:%S") if obj.started_at else "",
                obj.completed_at.strftime("%Y-%m-%d %H:%M:%S") if obj.completed_at else "",
                obj.user.email,
                obj.user.username,
                getattr(obj.user, 'dni', ''),
                obj.score_obtained,
                obj.attempt_number,
            ]
            
            # Answer info for each question
            answers_map = {ans.question_id: ans for ans in AttemptAnswer.objects.filter(attempt=obj)}
            
            for q in questions:
                ans_obj = answers_map.get(q.id)
                q_options = options_map.get(q.id, [])
                
                if ans_obj:
                    # Format selected answer text with letters
                    selected_text = ""
                    if q.question_type == 'multiple_choice':
                        selected_parts = []
                        sel_opts = set(ans_obj.selected_options.all().values_list('id', flat=True))
                        for i, opt in enumerate(q_options):
                            if opt.id in sel_opts:
                                selected_parts.append(f"{chr(65+i)}. {opt.text}")
                        selected_text = ", ".join(selected_parts)
                    elif q.question_type == 'open_ended':
                        selected_text = ans_obj.text_response or ""
                    else:
                        if ans_obj.selected_option:
                            index = -1
                            for i, opt in enumerate(q_options):
                                if opt.id == ans_obj.selected_option_id:
                                    index = i
                                    break
                            if index != -1:
                                selected_text = f"{chr(65+index)}. {ans_obj.selected_option.text}"
                            else:
                                selected_text = ans_obj.selected_option.text
                        else:
                            selected_text = ""
                    
                    row.append(selected_text)
                    row.append(ans_obj.points_obtained)
                else:
                    # Question not in this attempt
                    row.append("")
                    row.append(0)
            
            writer.writerow(row)
            
        return response

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def questions(self, request, pk=None):
        exam = self.get_object()
        user = request.user
        
        # Get or create an active attempt
        attempt = Attempt.objects.filter(user=user, exam=exam, status='in_progress').first()
        
        if not attempt:
            # Smart selection logic:
            
            # 1. Get IDs of questions user has EVER answered in this exam
            answered_question_ids = AttemptAnswer.objects.filter(
                attempt__user=user,
                attempt__exam=exam,
                attempt__status='completed'
            ).values_list('question_id', flat=True).distinct()
            
            # 2. Get IDs of questions user has EVER answered CORRECTLY
            correct_question_ids = AttemptAnswer.objects.filter(
                attempt__user=user,
                attempt__exam=exam,
                attempt__status='completed',
                is_correct=True
            ).values_list('question_id', flat=True).distinct()
            
            all_questions = list(Question.objects.filter(exam=exam))
            
            # 3. Categorize questions
            unseen_questions = [q for q in all_questions if q.id not in answered_question_ids]
            failed_questions = [q for q in all_questions if q.id in answered_question_ids and q.id not in correct_question_ids]
            mastered_questions = [q for q in all_questions if q.id in correct_question_ids]
            
            num_needed = min(len(all_questions), exam.questions_per_attempt)
            selected_questions = []

            # Priority 1: Unseen questions
            if len(unseen_questions) > 0:
                take = min(len(unseen_questions), num_needed)
                selected_questions.extend(random.sample(unseen_questions, take))
            
            # Priority 2: Failed questions (if still needed)
            needed_still = num_needed - len(selected_questions)
            if needed_still > 0 and len(failed_questions) > 0:
                take = min(len(failed_questions), needed_still)
                selected_questions.extend(random.sample(failed_questions, take))
            
            # Priority 3: Mastered questions (to fill the 10, if everything else is exhausted)
            needed_still = num_needed - len(selected_questions)
            if needed_still > 0 and len(mastered_questions) > 0:
                take = min(len(mastered_questions), needed_still)
                selected_questions.extend(random.sample(mastered_questions, take))
            
            # Shuffle the final selection so they don't appear in "priority order"
            random.shuffle(selected_questions)
            
            attempt = Attempt.objects.create(user=user, exam=exam)
            
            # Create AttemptQuestions to preserve order
            for i, q in enumerate(selected_questions):
                AttemptQuestion.objects.create(attempt=attempt, question=q, order_number=i+1)
        
        # Return questions in the attempt
        questions = [aq.question for aq in attempt.attempt_questions.all()]
        serializer = QuestionSerializer(questions, many=True)
        return Response({
            'attempt_id': attempt.id,
            'questions': serializer.data
        })

    @action(detail=False, methods=['get'], permission_classes=[IsStaff])
    def stats_summary(self, request):
        from django.db.models import Avg, Count, Max, Q, OuterRef, Subquery, IntegerField
        from django.contrib.auth import get_user_model
        
        UserModel = get_user_model()
        
        # Get all exams
        exams = Exam.objects.all()
        stats = []
        
        for exam in exams:
            # Subquery to find the best score for each analyst for this specific exam
            best_attempt_sq = Attempt.objects.filter(
                user=OuterRef('pk'),
                exam=exam,
                status='completed',
                counts_for_score=True
            ).order_by('-score_obtained').values('score_obtained')[:1]
            
            # Subquery for the ID of that best attempt (to use for question stats)
            best_attempt_id_sq = Attempt.objects.filter(
                user=OuterRef('pk'),
                exam=exam,
                status='completed',
                counts_for_score=True
            ).order_by('-score_obtained').values('id')[:1]
            
            # Annotate analysts with their best score for this exam
            analysts_with_scores = UserModel.objects.filter(is_staff=False).annotate(
                best_exam_score=Subquery(best_attempt_sq, output_field=IntegerField()),
                best_attempt_id=Subquery(best_attempt_id_sq, output_field=IntegerField())
            ).filter(best_exam_score__isnull=False)
            
            # Aggregate stats
            agg = analysts_with_scores.aggregate(
                total_analysts=Count('id'),
                avg_score=Avg('best_exam_score')
            )
            
            total_analysts = agg['total_analysts']
            avg_score = agg['avg_score'] or 0
            
            # Get best attempt IDs for question analysis
            best_attempt_ids = list(analysts_with_scores.values_list('best_attempt_id', flat=True))
            
            # Question stats logic
            q_stats = []
            if best_attempt_ids:
                questions = Question.objects.filter(exam=exam).prefetch_related('options')
                
                # Batch count answers to avoid N+1
                all_relevant_answers = AttemptAnswer.objects.filter(
                    question__exam=exam,
                    attempt_id__in=best_attempt_ids
                ).select_related('question')
                
                # Pre-calculate distribution
                for q in questions:
                    q_answers = [a for a in all_relevant_answers if a.question_id == q.id]
                    total_q = len(q_answers)
                    correct_q = len([a for a in q_answers if a.is_correct])
                    
                    choices = []
                    for opt in q.options.all():
                        if q.question_type == 'multiple_choice':
                            count = AttemptAnswer.objects.filter(id__in=[a.id for a in q_answers], selected_options=opt).count()
                        else:
                            count = len([a for a in q_answers if a.selected_option_id == opt.id])
                        
                        choices.append({
                            'text': opt.text,
                            'is_correct': opt.is_correct,
                            'count': count,
                            'percent': round((count / total_q * 100) if total_q > 0 else 0)
                        })

                    q_stats.append({
                        'id': q.id,
                        'text': q.text,
                        'type': q.question_type,
                        'total': total_q,
                        'correct': correct_q,
                        'percent': round((correct_q / total_q * 100) if total_q > 0 else 0),
                        'choices': choices
                    })

            stats.append({
                'id': exam.id,
                'name': exam.name,
                'total_attempts': total_analysts, 
                'avg_score': round(avg_score),
                'question_stats': q_stats
            })
            
        return Response(stats)

    @action(detail=True, methods=['get'], permission_classes=[IsStaff])
    def all_questions(self, request, pk=None):
        """
        Retorna todas las preguntas y opciones de un examen (para vista previa de administrador).
        No crea intentos ni limita el número de preguntas.
        """
        exam = self.get_object()
        questions = Question.objects.filter(exam=exam).order_by('id')
        serializer = QuestionSerializer(questions, many=True)
        return Response({
            'exam_name': exam.name,
            'total_questions': questions.count(),
            'questions': serializer.data
        })

class AttemptViewSet(viewsets.ModelViewSet):
    queryset = Attempt.objects.all()
    serializer_class = AttemptSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return Attempt.objects.all()
        return Attempt.objects.filter(user=self.request.user)

    @action(detail=False, methods=['get'])
    def user_results(self, request):
        if not request.user.is_staff:
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        from django.contrib.auth import get_user_model
        from django.db.models import Prefetch
        UserModel = get_user_model()
        
        search_query = request.query_params.get('search', '')
        offset = int(request.query_params.get('offset', 0))
        limit = 10
        
        analysts = UserModel.objects.filter(is_staff=False)
        if search_query:
            # Optimize: use username__icontains but ensure it's indexed
            analysts = analysts.filter(username__icontains=search_query)
            
        total_count = analysts.count()
        # Optimize: avoid loading all users if not needed
        analysts_page = analysts.order_by('username')[offset:offset+limit]
        
        # Optimize: Prefetch attempts and their answers to avoid N+1
        attempts_prefetch = Prefetch(
            'attempts',
            queryset=Attempt.objects.filter(status='completed').select_related('exam').prefetch_related(
                Prefetch('answers', queryset=AttemptAnswer.objects.select_related('question').prefetch_related('selected_options'))
            ).order_by('-completed_at'),
            to_attr='completed_attempts'
        )
        
        analysts_with_data = analysts_page.prefetch_related(attempts_prefetch)
        
        results = []
        for user in analysts_with_data:
            attempts_data = []
            for attempt in user.completed_attempts:
                ans_data = []
                for a in attempt.answers.all():
                    selected_text = "N/A"
                    if a.question.question_type == 'multiple_choice':
                        selected_text = ", ".join([o.text for o in a.selected_options.all()])
                    elif a.question.question_type == 'open_ended':
                        selected_text = a.text_response or ""
                    else:
                        selected_text = a.selected_option.text if a.selected_option else "N/A"

                    ans_data.append({
                        'question': a.question.text,
                        'selected': selected_text,
                        'is_correct': a.is_correct
                    })
                
                attempts_data.append({
                    'id': attempt.id,
                    'exam_id': attempt.exam.id,
                    'exam_name': attempt.exam.name,
                    'score': attempt.score_obtained,
                    'date': attempt.completed_at,
                    'attempt_number': attempt.attempt_number,
                    'counts_for_score': attempt.counts_for_score,
                    'answers': ans_data
                })
            
            results.append({
                'id': user.id,
                'username': user.username,
                'total_score': user.total_score,
                'attempts': attempts_data
            })
            
        return Response({
            'results': results,
            'total': total_count,
            'has_more': (offset + limit) < total_count
        })

    @action(detail=True, methods=['post'])
    def submit_answers(self, request, pk=None):
        attempt = self.get_object()
        if attempt.status != 'in_progress':
            return Response({'error': 'Attempt is already finished'}, status=status.HTTP_400_BAD_REQUEST)
        
        answers_data = request.data.get('answers', [])
        total_points = 0
        
        for ans in answers_data:
            question_id = ans.get('question_id')
            option_id = ans.get('selected_option_id') # single
            option_ids = ans.get('selected_option_ids', []) # multiple
            text_response = ans.get('text_response') # open
            
            try:
                question = Question.objects.get(id=question_id, exam=attempt.exam)
                
                # Get or create answer object
                answer_obj, created = AttemptAnswer.objects.get_or_create(
                    attempt=attempt,
                    question=question
                )
                
                if question.question_type == 'single_choice':
                    option = Option.objects.get(id=option_id, question=question) if option_id else None
                    answer_obj.selected_option = option
                    answer_obj.save() # calculates points in models.py
                    
                elif question.question_type == 'multiple_choice':
                    # Multiple choice scoring: all correct options must be selected, no incorrect ones
                    options = Option.objects.filter(id__in=option_ids, question=question)
                    answer_obj.selected_options.set(options)
                    
                    correct_options_ids = set(Option.objects.filter(question=question, is_correct=True).values_list('id', flat=True))
                    selected_ids = set(option_ids)
                    
                    if selected_ids == correct_options_ids and len(correct_options_ids) > 0:
                        answer_obj.is_correct = True
                        answer_obj.points_obtained = question.points
                    else:
                        answer_obj.is_correct = False
                        answer_obj.points_obtained = 0
                    answer_obj.save()
                    
                elif question.question_type == 'open_ended':
                    answer_obj.text_response = text_response
                    answer_obj.save() # calculates points in models.py (non-empty = correct)
                
                total_points += answer_obj.points_obtained
            except (Question.DoesNotExist, Option.DoesNotExist):
                continue
        
        # Calculate final score (0-100)
        aq_count = AttemptQuestion.objects.filter(attempt=attempt).count()
        max_possible_points = AttemptQuestion.objects.filter(attempt=attempt).aggregate(total=models.Sum('question__points'))['total'] or 0
        
        if max_possible_points > 0:
            score = round((total_points / max_possible_points) * 100)
        else:
            score = 0
            
        attempt.status = 'completed'
        attempt.score_obtained = score
        attempt.completed_at = timezone.now()
        attempt.save()
        
        # Update user total_score if valid
        if attempt.counts_for_score:
            user = attempt.user
            user.total_score += score
            user.save()
            user.update_rank()
            
        # Calculate correct answers count
        correct_count = AttemptAnswer.objects.filter(attempt=attempt, is_correct=True).count()
        total_questions = AttemptQuestion.objects.filter(attempt=attempt).count()
        
        # Calculate attempts left for scoring
        attempts_count = Attempt.objects.filter(user=attempt.user, exam=attempt.exam).count()
        attempts_left = max(0, attempt.exam.max_scored_attempts - attempts_count)

        return Response({
            'score': score,
            'correct_count': correct_count,
            'total_questions': total_questions,
            'status': attempt.status,
            'counts_for_score': attempt.counts_for_score,
            'attempts_left': attempts_left,
            'total_user_score': attempt.user.total_score
        })

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsStaff]
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        exam_id = self.request.query_params.get('exam_id')
        if exam_id:
            return Question.objects.filter(exam_id=exam_id).order_by('id')
        return Question.objects.all()

class OptionViewSet(viewsets.ModelViewSet):
    from .serializers import OptionSerializer
    queryset = Option.objects.all()
    serializer_class = OptionSerializer
    permission_classes = [IsStaff]

    def get_queryset(self):
        question_id = self.request.query_params.get('question_id')
        if question_id:
            return Option.objects.filter(question_id=question_id).order_by('id')
        return Option.objects.all()
